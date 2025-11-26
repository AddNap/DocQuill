"""
XLSX exporter for DOCX documents.

Exports document content to XLSX format using openpyxl.
"""

from typing import Dict, Any, List, Optional, Union
from pathlib import Path
import logging
from .base_exporter import BaseExporter

logger = logging.getLogger(__name__)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl not available - XLSX export will not work")


class XLSXExporter(BaseExporter):
    """
    Exports DOCX documents to XLSX format.
    """
    
    def __init__(self, document, output_path: Optional[str] = None, 
                 export_options: Optional[Dict[str, Any]] = None):
        """
        Initialize XLSX exporter.
        
        Args:
            document: Document to export
            output_path: Output path for XLSX file
            export_options: Export options
        """
        super().__init__(document, output_path, export_options)
        
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for XLSX export")
        
        # XLSX-specific options
        self.include_headers = self.get_export_option('include_headers', True)
        self.export_tables_only = self.get_export_option('export_tables_only', False)
        self.export_paragraphs_as_rows = self.get_export_option('export_paragraphs_as_rows', True)
        self.include_metadata = self.get_export_option('include_metadata', False)
        self.apply_formatting = self.get_export_option('apply_formatting', True)
        self.auto_adjust_columns = self.get_export_option('auto_adjust_columns', True)
        self.sheet_name = self.get_export_option('sheet_name', 'Document Content')
        
        # Ensure options are in export_options for validation
        self.export_options.setdefault('include_headers', self.include_headers)
        self.export_options.setdefault('export_tables_only', self.export_tables_only)
        self.export_options.setdefault('export_paragraphs_as_rows', self.export_paragraphs_as_rows)
        self.export_options.setdefault('include_metadata', self.include_metadata)
        self.export_options.setdefault('apply_formatting', self.apply_formatting)
        self.export_options.setdefault('auto_adjust_columns', self.auto_adjust_columns)
        self.export_options.setdefault('sheet_name', self.sheet_name)
    
    def export_to_file(self, file_path: Optional[str] = None) -> bool:
        """
        Export document to XLSX file.
        
        Args:
            file_path: Output file path (uses output_path if not provided)
            
        Returns:
            True if successful, False otherwise
        """
        try:
            if file_path is None:
                file_path = self.output_path
            
            if file_path is None:
                raise ValueError("No output path specified")
            
            # Convert to Path if needed
            if isinstance(file_path, str):
                file_path = Path(file_path)
            
            # Create directory if it doesn't exist
            file_path.parent.mkdir(parents=True, exist_ok=True)
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = self.sheet_name
            
            # Export content
            if self.export_tables_only:
                self._export_tables_only(ws)
            else:
                self._export_full_content(ws)
            
            # Auto-adjust column widths
            if self.auto_adjust_columns:
                self._auto_adjust_columns(ws)
            
            # Save workbook
            wb.save(file_path)
            
            logger.info(f"XLSX exported to {file_path}")
            return True
            
        except Exception as e:
            logger.error(f"Failed to export to XLSX file: {e}")
            raise
    
    def _export_tables_only(self, ws):
        """Export only tables from the document."""
        if hasattr(self.document, 'get_tables'):
            tables = self.document.get_tables()
            
            current_row = 1
            
            for i, table in enumerate(tables):
                # Add table header
                ws.cell(row=current_row, column=1, value=f'Table {i + 1}')
                if self.apply_formatting:
                    ws.cell(row=current_row, column=1).font = Font(bold=True)
                current_row += 1
                
                # Export table content
                current_row = self._export_table(ws, table, current_row)
                
                # Add empty row between tables
                if i < len(tables) - 1:
                    current_row += 1
    
    def _export_full_content(self, ws):
        """Export full document content."""
        if hasattr(self.document, 'get_body'):
            body = self.document.get_body()
            if body:
                self._export_body(ws, body)
    
    def _export_body(self, ws, body, start_row: int = 1):
        """Export body content."""
        current_row = start_row
        
        for child in body.children:
            if hasattr(child, 'get_rows'):  # Table
                current_row = self._export_table(ws, child, current_row)
            elif self.export_paragraphs_as_rows and hasattr(child, 'get_text'):  # Paragraph
                text = child.get_text()
                if text.strip():
                    ws.cell(row=current_row, column=1, value='Paragraph')
                    ws.cell(row=current_row, column=2, value=text)
                    current_row += 1
            elif hasattr(child, 'get_text'):  # Other text elements
                text = child.get_text()
                if text.strip():
                    ws.cell(row=current_row, column=1, value=type(child).__name__)
                    ws.cell(row=current_row, column=2, value=text)
                    current_row += 1
        
        return current_row
    
    def _export_table(self, ws, table, start_row: int = 1):
        """Export table to XLSX."""
        if not hasattr(table, 'get_rows'):
            return start_row
        
        rows = table.get_rows()
        if not rows:
            return start_row
        
        current_row = start_row
        
        # Export header row if it exists
        if self.include_headers and rows:
            first_row = rows[0]
            if hasattr(first_row, 'is_header_row') and first_row.is_header_row:
                for col_idx, cell in enumerate(first_row.cells, 1):
                    ws.cell(row=current_row, column=col_idx, value=self._get_cell_text(cell))
                    if self.apply_formatting:
                        ws.cell(row=current_row, column=col_idx).font = Font(bold=True)
                        ws.cell(row=current_row, column=col_idx).fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
                current_row += 1
        
        # Export data rows
        for row in rows:
            if hasattr(row, 'is_header_row') and row.is_header_row:
                continue  # Skip header row if already exported
            
            for col_idx, cell in enumerate(row.cells, 1):
                ws.cell(row=current_row, column=col_idx, value=self._get_cell_text(cell))
            current_row += 1
        
        return current_row
    
    def _get_cell_text(self, cell) -> str:
        """Get text content from cell."""
        if hasattr(cell, 'get_text'):
            return cell.get_text()
        elif hasattr(cell, 'text'):
            return cell.text
        else:
            return str(cell)
    
    def _auto_adjust_columns(self, ws):
        """Auto-adjust column widths."""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def get_supported_formats(self) -> List[str]:
        """Get supported export formats."""
        return ['xlsx', 'xls']
    
    def get_export_info(self) -> Dict[str, Any]:
        """Get export information."""
        return {
            'format': 'XLSX',
            'include_headers': self.include_headers,
            'export_tables_only': self.export_tables_only,
            'export_paragraphs_as_rows': self.export_paragraphs_as_rows,
            'include_metadata': self.include_metadata,
            'apply_formatting': self.apply_formatting,
            'auto_adjust_columns': self.auto_adjust_columns,
            'sheet_name': self.sheet_name
        }
